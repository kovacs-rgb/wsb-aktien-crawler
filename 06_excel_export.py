"""
06 - Excel-Export
Liest alle Pickle-Ergebnisse und erstellt eine uebersichtliche Excel-Tabelle
mit bedingter Formatierung (3-Farben-Skala: rot-gelb-gruen).
"""
import os
import pickle

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PICKLE_DIR = os.path.join(BASE_DIR, "pickle")
OUTPUT_DIR = os.path.join(BASE_DIR, "output")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "crawler_results.xlsx")


def read_pickle_files(directory):
    """Liest alle Pickle-Dateien aus dem Verzeichnis."""
    data_list = []
    for filename in sorted(os.listdir(directory)):
        if filename.endswith(".pkl") or filename.endswith(".pickle"):
            filepath = os.path.join(directory, filename)
            try:
                with open(filepath, "rb") as f:
                    data = pickle.load(f)
                data_list.append(data)
                print(f"  Geladen: {filename}")
            except Exception as e:
                print(f"  FEHLER bei {filename}: {e}")
    return data_list


def create_dataframe(data_list):
    """Erstellt DataFrame mit allen Akronymen als Spalten."""
    all_acronyms = set()
    for entry in data_list:
        all_acronyms.update(entry["results"].keys())
    all_acronyms = sorted(all_acronyms)

    rows = []
    for entry in data_list:
        row = {acronym: 0 for acronym in all_acronyms}
        row.update(entry["results"])
        row["Run-ID"] = entry["run_id"]
        rows.append(row)

    df = pd.DataFrame(rows, columns=["Run-ID"] + all_acronyms)

    # Nach Run-ID sortieren (Datum+Uhrzeit)
    df = df.sort_values(by="Run-ID").reset_index(drop=True)
    return df


def write_excel_with_formatting(df, output_path):
    """Schreibt DataFrame in Excel mit bedingter 3-Farben-Formatierung."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Crawler-Ergebnisse"

    # Header schreiben
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = cell.font.copy(bold=True)

    # Daten schreiben
    for row_idx, (_, row_data) in enumerate(df.iterrows(), 2):
        for col_idx, value in enumerate(row_data.values, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Bedingte Formatierung: 3-Farben-Skala fuer Datenspalten (ab Spalte 2)
    if len(df) > 0:
        max_row = len(df) + 1
        for col_idx in range(2, len(df.columns) + 1):
            col_letter = get_column_letter(col_idx)
            cell_range = f"{col_letter}2:{col_letter}{max_row}"
            rule = ColorScaleRule(
                start_type="num", start_value=0, start_color="F8696B",  # Rot
                mid_type="percentile", mid_value=50, mid_color="FFEB84",  # Gelb
                end_type="max", end_color="63BE7B",  # Gruen
            )
            ws.conditional_formatting.add(cell_range, rule)

    # Spaltenbreite anpassen
    ws.column_dimensions["A"].width = 14
    for col_idx in range(2, len(df.columns) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 8

    wb.save(output_path)


def main():
    print("=== Excel-Export ===")

    if not os.path.exists(PICKLE_DIR):
        print(f"FEHLER: Pickle-Verzeichnis nicht gefunden: {PICKLE_DIR}")
        print("Bitte zuerst den Crawler (04_crawler.py) ausfuehren.")
        return

    # Pickle-Dateien laden
    print("Lade Pickle-Dateien...")
    data_list = read_pickle_files(PICKLE_DIR)

    if not data_list:
        print("Keine Pickle-Dateien gefunden!")
        return

    print(f"\n{len(data_list)} Datensaetze geladen.")

    # DataFrame erstellen
    df = create_dataframe(data_list)
    print(f"Tabelle: {df.shape[0]} Zeilen, {df.shape[1]} Spalten")

    # Excel schreiben
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    write_excel_with_formatting(df, OUTPUT_FILE)
    print(f"\nExcel-Datei gespeichert: {OUTPUT_FILE}")

    # Vorschau
    print(f"\nVorschau:")
    print(df.to_string(max_rows=10, max_cols=15))

    print(f"\nZeitraum: {df['Run-ID'].iloc[0]} bis {df['Run-ID'].iloc[-1]}")
    print(f"Anzahl Symbole: {len(df.columns) - 1}")


if __name__ == "__main__":
    main()
